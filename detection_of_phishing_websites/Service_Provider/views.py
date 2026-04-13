import json
from functools import wraps
from django.shortcuts import render, redirect
from django.http import HttpResponse, JsonResponse
from django.contrib import messages
from django.contrib.auth.models import User
from django.db.models import Count
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

from Remote_User.models import URLPrediction, ModelAccuracy, UserProfile

ADMIN_USERNAME = 'Admin'
ADMIN_PASSWORD = 'Admin'


# ─────────────────────────────────────────────────────────
# Admin auth decorator
# ─────────────────────────────────────────────────────────

def admin_required(view_func):
    @wraps(view_func)
    def wrapper(request, *args, **kwargs):
        if not request.session.get('is_admin'):
            return redirect('admin_login')
        return view_func(request, *args, **kwargs)
    return wrapper


# ─────────────────────────────────────────────────────────
# Auth views
# ─────────────────────────────────────────────────────────

def admin_login(request):
    if request.session.get('is_admin'):
        return redirect('admin_dashboard')

    if request.method == 'POST':
        username = request.POST.get('username', '')
        password = request.POST.get('password', '')
        if username == ADMIN_USERNAME and password == ADMIN_PASSWORD:
            request.session['is_admin'] = True
            return redirect('admin_dashboard')
        messages.error(request, 'Invalid admin credentials.')

    return render(request, 'SProvider/login.html')


def admin_logout(request):
    request.session.pop('is_admin', None)
    return redirect('admin_login')


# ─────────────────────────────────────────────────────────
# Dashboard
# ─────────────────────────────────────────────────────────

@admin_required
def admin_dashboard(request):
    total_users = User.objects.filter(is_staff=False).count()
    total_predictions = URLPrediction.objects.count()
    phishing_count = URLPrediction.objects.filter(result='Phishing').count()
    safe_count = URLPrediction.objects.filter(result='Non Phishing').count()
    recent_predictions = URLPrediction.objects.select_related('user').order_by('-checked_at')[:8]
    phishing_pct = round(phishing_count / total_predictions * 100, 1) if total_predictions else 0

    return render(request, 'SProvider/dashboard.html', {
        'total_users': total_users,
        'total_predictions': total_predictions,
        'phishing_count': phishing_count,
        'safe_count': safe_count,
        'phishing_pct': phishing_pct,
        'recent_predictions': recent_predictions,
    })


# ─────────────────────────────────────────────────────────
# Users
# ─────────────────────────────────────────────────────────

@admin_required
def view_users(request):
    users = User.objects.filter(is_staff=False).order_by('-date_joined').prefetch_related('predictions')
    user_data = []
    for user in users:
        total = user.predictions.count()
        phishing = user.predictions.filter(result='Phishing').count()
        user_data.append({
            'user': user,
            'total': total,
            'phishing': phishing,
            'safe': total - phishing,
        })
    return render(request, 'SProvider/users.html', {'user_data': user_data})


# ─────────────────────────────────────────────────────────
# ML Training
# ─────────────────────────────────────────────────────────

@admin_required
def train_model(request):
    if request.method == 'POST':
        try:
            from ml_engine import train_all_models
            accuracy_results, _, _ = train_all_models()

            ModelAccuracy.objects.all().delete()
            for model_name, accuracy in accuracy_results.items():
                ModelAccuracy.objects.create(model_name=model_name, accuracy=accuracy)

            return JsonResponse({
                'success': True,
                'results': accuracy_results,
                'message': f'Successfully trained {len(accuracy_results)} models.',
            })
        except FileNotFoundError:
            return JsonResponse({'error': 'Dataset CSV file not found.'}, status=500)
        except Exception as e:
            return JsonResponse({'error': str(e)}, status=500)

    accuracies = ModelAccuracy.objects.all()
    last_trained = accuracies.first().trained_at if accuracies.exists() else None

    return render(request, 'SProvider/train.html', {
        'accuracies': accuracies,
        'last_trained': last_trained,
    })


# ─────────────────────────────────────────────────────────
# Predictions
# ─────────────────────────────────────────────────────────

@admin_required
def view_predictions(request):
    filter_type = request.GET.get('filter', 'all')
    predictions = URLPrediction.objects.select_related('user').order_by('-checked_at')

    if filter_type == 'phishing':
        predictions = predictions.filter(result='Phishing')
    elif filter_type == 'safe':
        predictions = predictions.filter(result='Non Phishing')

    return render(request, 'SProvider/predictions.html', {
        'predictions': predictions,
        'filter_type': filter_type,
        'total': predictions.count(),
    })


# ─────────────────────────────────────────────────────────
# Analytics
# ─────────────────────────────────────────────────────────

@admin_required
def analytics(request):
    total = URLPrediction.objects.count()
    phishing = URLPrediction.objects.filter(result='Phishing').count()
    safe = URLPrediction.objects.filter(result='Non Phishing').count()
    accuracies = list(ModelAccuracy.objects.values('model_name', 'accuracy'))

    # Top users by predictions
    top_users = (
        User.objects.filter(is_staff=False)
        .annotate(pred_count=Count('predictions'))
        .order_by('-pred_count')[:5]
    )

    return render(request, 'SProvider/analytics.html', {
        'total': total,
        'phishing': phishing,
        'safe': safe,
        'accuracies': accuracies,
        'accuracies_json': json.dumps(accuracies),
        'phishing_ratio': round(phishing / total * 100, 1) if total else 0,
        'top_users': top_users,
    })


# ─────────────────────────────────────────────────────────
# Export
# ─────────────────────────────────────────────────────────

@admin_required
def download_excel(request):
    wb = Workbook()
    ws = wb.active
    ws.title = 'Predictions'

    # Header row styling
    header_font = Font(bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='1e3a5f', end_color='1e3a5f', fill_type='solid')
    headers = ['#', 'URL', 'Result', 'User', 'Email', 'Date Checked']

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')

    # Data rows
    predictions = URLPrediction.objects.select_related('user').order_by('-checked_at')
    for row, pred in enumerate(predictions, 2):
        ws.cell(row=row, column=1, value=row - 1)
        ws.cell(row=row, column=2, value=pred.url)
        ws.cell(row=row, column=3, value=pred.result)
        ws.cell(row=row, column=4, value=pred.user.username)
        ws.cell(row=row, column=5, value=pred.user.email)
        ws.cell(row=row, column=6, value=pred.checked_at.strftime('%Y-%m-%d %H:%M'))

    # Column widths
    ws.column_dimensions['A'].width = 6
    ws.column_dimensions['B'].width = 60
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['D'].width = 20
    ws.column_dimensions['E'].width = 30
    ws.column_dimensions['F'].width = 20

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename="PhishGuard_Predictions.xlsx"'
    wb.save(response)
    return response
